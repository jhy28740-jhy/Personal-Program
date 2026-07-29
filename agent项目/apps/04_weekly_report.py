"""
周报自动生成器（实战项目）
=====================================
功能：
1. 扫描"本周"工作文件夹，收集每个任务的结构和脚本内容作为证据
2. 用 Agent 智能总结成专业的"工作进展"描述
3. 填入周报模板的"工作进展"列，另存为新文件（原模板不动）

用法：
    cd "C:\\Users\\hyji11\\Desktop\\个人小项目\\agent项目"
    agent_env\\Scripts\\activate
    python 04_weekly_report.py
"""

import os
import sys
import io
import shutil
from pathlib import Path
from datetime import datetime, timedelta
from dotenv import load_dotenv
from crewai import Agent, Task, Crew, LLM

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

load_dotenv()

# ============ 配置区（可按需修改）============
WORK_DIR = Path(r"C:\Users\hyji11\Desktop\端到端脚本-常用\工作任务\七月")
TEMPLATE = Path(r"C:\Users\hyji11\Downloads\钱瑶个人周报(AI测试部).xlsx")
OUTPUT_DIR = Path(r"C:\Users\hyji11\Desktop\个人小项目\agent项目\周报输出")

# 本周日期范围（默认：本周一到周日）
TODAY = datetime.now()
WEEK_START = TODAY - timedelta(days=TODAY.weekday())        # 本周一
WEEK_END = WEEK_START + timedelta(days=6)                    # 本周日

# 读取脚本内容的文件类型（这些最能说明"做了什么"）
CODE_EXTS = {".py", ".txt", ".md", ".json", ".sh", ".sql"}
MAX_FILE_CHARS = 1500   # 每个脚本文件最多读多少字符

# ============ LLM 配置 ============
llm = LLM(
    model="openai/gpt-5.6-sol",
    api_key=os.getenv("OPENAI_API_KEY"),
    base_url=os.getenv("OPENAI_API_BASE")
)


def parse_folder_date(name: str):
    """从文件夹名解析日期，如 '7月20日氢离子脚本跑测' -> date(2026,7,20)。解析失败返回None。"""
    import re
    m = re.search(r"(\d{1,2})月(\d{1,2})[日号]", name)
    if not m:
        return None
    month, day = int(m.group(1)), int(m.group(2))
    try:
        return datetime(TODAY.year, month, day)
    except ValueError:
        return None


def find_this_week_folders():
    """找出本周的工作文件夹，按日期排序。"""
    result = []
    for item in WORK_DIR.iterdir():
        if not item.is_dir():
            continue
        d = parse_folder_date(item.name)
        if d and WEEK_START.date() <= d.date() <= WEEK_END.date():
            result.append((d, item))
    result.sort(key=lambda x: x[0])
    return [item for _, item in result]


def collect_evidence(folder: Path) -> str:
    """收集一个文件夹的证据：目录树 + 小脚本文件内容。"""
    lines = [f"【任务文件夹】{folder.name}", "目录结构："]
    code_files = []
    for p in sorted(folder.rglob("*")):
        rel = p.relative_to(folder)
        depth = len(rel.parts) - 1
        indent = "  " * depth
        if p.is_dir():
            lines.append(f"{indent}📁 {p.name}/")
        else:
            size_kb = p.stat().st_size / 1024
            lines.append(f"{indent}📄 {p.name} ({size_kb:.0f}KB)")
            if p.suffix.lower() in CODE_EXTS and p.stat().st_size < 50000:
                code_files.append(p)
    # 附上脚本内容
    for p in code_files[:5]:   # 最多读5个脚本
        try:
            text = p.read_text(encoding="utf-8", errors="ignore")[:MAX_FILE_CHARS]
            lines.append(f"\n--- 脚本内容：{p.name} ---\n{text}")
        except Exception:
            pass
    return "\n".join(lines)


# ============ 创建总结 Agent ============
summarizer = Agent(
    role='测试工程师周报助手',
    goal='根据工作文件夹的证据，总结出专业、精炼的工作进展描述',
    backstory=(
        '你是一位AI测试部门的资深工程师，擅长把零散的工作痕迹（脚本、数据、文件）'
        '提炼成一句专业、准确的工作进展。你的总结面向周报，要简洁（30字以内），'
        '突出"做了什么、达到什么效果"，使用测试/算法领域的专业表达。'
    ),
    verbose=False,
    llm=llm
)


def summarize_work(evidence: str) -> str:
    """让Agent根据证据总结一句工作进展。"""
    task = Task(
        description=(
            f"以下是我本周某个工作任务的文件夹证据，请总结成一句专业的工作进展描述"
            f"（30字以内，面向周报，突出做了什么和效果）：\n\n{evidence}"
        ),
        expected_output='一句30字以内的专业工作进展描述，不要加引号或多余说明',
        agent=summarizer
    )
    crew = Crew(agents=[summarizer], tasks=[task], verbose=False)
    return str(crew.kickoff()).strip()


def write_report(summaries: list):
    """把总结写入模板C列，另存为新文件。summaries是[(文件夹名, 进展)]列表。"""
    import openpyxl
    OUTPUT_DIR.mkdir(exist_ok=True)
    out_name = f"纪浩阳个人周报(AI测试部)_{WEEK_START:%m%d}-{WEEK_END:%m%d}.xlsx"
    out_path = OUTPUT_DIR / out_name
    shutil.copy(TEMPLATE, out_path)   # 复制模板，保护原文件

    wb = openpyxl.load_workbook(out_path)
    ws = wb["周报"]

    # 更新标题日期（第2行）
    ws["A2"] = f"纪浩阳个人周报(AI测试部)({WEEK_START:%Y%m%d}-{WEEK_END:%Y%m%d})"

    # 保持模板"一个格子里分行"的布局，只操作第7行
    # C列：填入带序号的工作进展（自动生成）
    progress_text = "\n".join(f"{i+1}、{s}" for i, (_, s) in enumerate(summaries))
    ws["C7"] = progress_text

    # D列（工作进度）：只留序号占位 "1、\n2、\n3、\n4、"，百分比你自己填
    order_placeholder = "\n".join(f"{i+1}、" for i in range(len(summaries)))
    ws["D7"] = order_placeholder

    # E/F/G 列（下周计划/依赖项/说明）：清空，留给你手填
    ws["E7"] = None
    ws["F7"] = None
    ws["G7"] = None

    wb.save(out_path)
    return out_path


def main():
    print("=" * 55)
    print(f"  📊 周报自动生成器")
    print(f"  本周范围：{WEEK_START:%Y-%m-%d} ~ {WEEK_END:%Y-%m-%d}")
    print("=" * 55)

    # 1. 找本周文件夹
    folders = find_this_week_folders()
    if not folders:
        print("⚠️ 没有找到本周的工作文件夹，请检查日期范围或路径。")
        return
    print(f"\n找到 {len(folders)} 个本周任务：")
    for f in folders:
        print(f"  • {f.name}")

    # 2. 逐个总结
    summaries = []
    for folder in folders:
        print(f"\n🔍 正在分析：{folder.name} ...")
        evidence = collect_evidence(folder)
        summary = summarize_work(evidence)
        print(f"   ✅ 进展：{summary}")
        summaries.append((folder.name, summary))

    # 3. 写入周报
    print("\n📝 正在生成周报文件...")
    out_path = write_report(summaries)
    print(f"\n{'=' * 55}")
    print(f"✅ 周报已生成：{out_path}")
    print(f"   （工作进展列已填充，工作进度/下周计划请你手动补充）")
    print("=" * 55)


if __name__ == "__main__":
    main()
