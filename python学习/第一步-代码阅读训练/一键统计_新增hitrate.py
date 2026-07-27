# -*- coding: utf-8 -*-
"""一键统计：三种模式（场景统计 / 类型统计 / 普通总计统计）—— 完全独立版

无任何外部依赖，所有统计逻辑内嵌；权重表直接硬编码进代码。
传给别人使用时只需要：
  - 这个脚本
  - 标准 Python + pandas + openpyxl 环境

入参：
  - N 个 A 表（带标注的 *_AB合并表.xlsx，含 query/title/content/relevance_label/...）
  - 1 个映射表
  - STATS_MODE：选 "场景" / "类型" / "普通总计"

各模式映射表要求：
  - "场景"     : rawquery, query, 一级场景, 二级场景, 三级场景
  - "类型"     : rawquery, query, 类型
  - "普通总计" : rawquery, query

各模式输出：
  - "场景"   ：多级场景汇总 + 二级场景对比 + 总计加权对比（共 3 张表）
  - "类型"   ：6 链路 × (全量+各类型) 维度统计总表（1 张）
  - "普通总计"：6 链路 × 全量 维度统计总表（1 张）

链路名：从 A 表文件名前缀自动提取（"_" 分隔取第一段）

使用方法：
    直接修改本文件 if __name__ == "__main__": 块下面的配置区，然后运行：
    python 一键统计_三模式_独立版.py
"""
import math
import os
import sys
import time
from collections import Counter
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ============================================================
# 硬编码场景权重（来自 测试集1584映射表(1).xlsx 线上分布 sheet，2026-06-17 更新）
# 改为按二级场景加权（原为一级场景）
# ============================================================
SCENE_WEIGHTS = {
    "诊疗-治疗": 0.237821,
    "诊疗-用药": 0.214825,
    "诊疗-检查化验": 0.208133,
    "诊疗-重症护理": 0.050317,
    "诊疗-饮食指导": 0.043402,
    "预防-疾病预防": 0.023646,
    "生育-生育": 0.020149,
    "衍生-医疗常识": 0.018002,
    "就医-就医信息": 0.015922,
    "诊疗-康复指导": 0.013824,
    "诊疗-重症预防": 0.010004,
    "保健-运动指导": 0.007673,
    "诊疗-病因": 0.007319,
    "保健-睡眠指导": 0.005567,
    "诊疗-医疗伦理": 0.004063,
    "诊疗-健康宣教": 0.003877,
    "预防-康复预防": 0.003849,
    "保健-婴幼护理": 0.003349,
    "诊疗-流行病学": 0.002374,
    "诊疗-康复指南": 0.002084,
    "衍生-器械设备": 0.001999,
    "诊疗-病程": 0.001874,
    "诊疗-症状": 0.001845,
    "其他-研究进展": 0.001693,
    "其他-学科职业": 0.001460,
    "预防-孕前备孕": 0.001320,
    "诊疗-产前产后": 0.000679,
    "诊疗-医学自然": 0.000459,
    "诊疗-医院见闻": 0.000222,
    "诊疗-政策法规": 0.000215,
    "诊疗-医学知识问答": 0.000040,
    "无分类": 0.091994,  # 一级场景为"无分类"且二级场景为"/"
}

# 时间戳合理范围（1970-2030）
TS_MIN = 0
TS_MAX = 1924992000

OVERALL_NAME = "总计(按照二级场景加权计算)"
DIFF_NAME = "总计diff"

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BOLD_FONT = Font(bold=True)

METRIC_COLS = ["有结果率", "有结果吸收率", "hit_rate", "产出率", "Agood率", "A+Bgood率",
               "query_Agood率", "query_A+Bgood率",
               "NDCG", "MAP", "MRR", "Timeliness", "Authority"]


# ============================================================
# 通用：标签映射 + 指标
# ============================================================
def map_relevance_label(value):
    """A/B/C/D → 4/3/2/1，弃标→0，其他→None"""
    if pd.isna(value) or value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        if v in ("弃标", ""):
            return 0
        m = {"A": 4, "B": 3, "C": 2, "D": 1}.get(v.upper())
        if m is not None:
            return m
        try:
            return int(v)
        except (ValueError, TypeError):
            return None
    if isinstance(value, (int, float)):
        return int(value)
    return None


def dcg(scores, k):
    v = 0.0
    for i in range(1, min(len(scores) + 1, k + 1)):
        if scores[i - 1] > 0:
            v += (2 ** scores[i - 1] - 1) / math.log2(i + 1)
    return v


def calc_mrr(scores):
    for rank, s in enumerate(scores, 1):
        if s in (3, 4):
            return 1.0 / rank
    return 0.0


def calc_map(scores):
    pos, prec = [], []
    for k, s in enumerate(scores, 1):
        if s in (3, 4):
            pos.append(k)
            prec.append(len(pos) / k)
    return sum(prec) / len(prec) if prec else 0


def ts_to_valid_date(ts):
    """过滤异常时间戳，返回有效日期字符串或 None"""
    try:
        v = int(ts)
        if v <= 0:
            return None
        if v > 1e12:
            v = v // 1000
        if v < TS_MIN or v > TS_MAX:
            return None
        return time.strftime("%Y-%m-%d", time.localtime(v))
    except (ValueError, TypeError, OverflowError):
        return None


def time_decay(date_str, rate=0.161):
    d = datetime.strptime(date_str, "%Y-%m-%d")
    yr = (datetime.now() - d).days / 365.0
    if yr < 0:
        yr = 0
    return math.exp(-rate * yr)


def auth_score(row):
    """权威度评分（0-6）"""
    idx = str(row.get("indexName", "") or row.get("_indexName", ""))
    auth = str(row.get("authority", ""))
    cat = str(row.get("category", ""))
    if "期刊" in idx:
        if auth in ["双核心", "单核心", "非核心"]:
            return {"双核心": 3, "单核心": 2, "非核心": 1}.get(auth, 0)
        core_count = auth.count("核心")
        if core_count >= 2:
            return 3
        elif core_count == 1:
            return 2
        else:
            return 1
    if "指南" in idx:
        return {"专家共识": 4, "指南": 6, "指南解读": 3, "规范": 5,
                "专家建议": 3, "标准": 5, "指南介绍": 3, "指南摘译": 3}.get(cat, 3)
    if "教科书" in idx or "书籍" in idx:
        return 6
    if "药品" in idx:
        return 6
    if "疾病" in idx:
        return 2
    return 0


def run_stats(df_full, query_set, qt, topk, raw_query_map=None):
    """对 query 集合算所有指标

    raw_query_map: dict[rawquery -> set(query)]，用于计算 hit_rate
                   - 分母：raw_query_map 的 key 数（去重 rawquery 数）
                   - 分子：rawquery 拆出的任一 query 在该链路返回了 A/B 级 doc，则该 rawquery hit
    """
    df = df_full[df_full["query"].isin(query_set)].copy()
    rel = "relevance_label"
    if rel not in df.columns:
        return None
    df[rel] = df[rel].apply(map_relevance_label)
    df = df[df[rel].notna()].copy()

    sort_col = "_rank_index" if "_rank_index" in df.columns else (
        "index" if "index" in df.columns else None)
    if sort_col:
        df = df.sort_values(["query", sort_col]).reset_index(drop=True)

    grouped = df.groupby("query", sort=False)
    ndcg_list, mrr_list, map_list = [], [], []
    doc_total = a_total = ab_total = 0
    qry_with_good = qry_count = 0
    qa_rates, qab_rates = [], []
    yield_sum = 0.0
    year_scores, auth_scores = [], []
    query_has_good = set()  # 该链路命中过 A/B 的 query 集合（用于 hit_rate）

    for _, g in grouped:
        s = g[rel].astype(int).tolist()
        if not s:
            continue
        qry_count += 1
        s_topk = s[:topk]
        dc = len(s_topk)
        gs = sorted(s, reverse=True)
        idcg_v = dcg(gs, topk)
        ndcg_list.append(dcg(s_topk, topk) / idcg_v if idcg_v > 0 else 0)
        mrr_list.append(calc_mrr(s_topk))
        m = calc_map(s_topk)
        if m > 0:
            map_list.append(m)
        doc_total += dc
        ac = Counter(s_topk)[4]
        bc = Counter(s_topk)[3]
        a_total += ac
        ab_total += (ac + bc)
        qa_rates.append(ac / dc)
        qab_rates.append((ac + bc) / dc)
        if (ac + bc) > 0:
            qry_with_good += 1
            query_has_good.add(g["query"].iloc[0])
        yield_sum += dc / topk
        for ts in g.get("post_ts", []):
            ds = ts_to_valid_date(ts)
            if ds:
                year_scores.append(time_decay(ds))
        for _, row in g.iterrows():
            auth_scores.append(auth_score(row))

    if qry_count == 0:
        return None

    # hit_rate：rawquery 级命中率
    hit_rate = 0.0
    if raw_query_map:
        raw_total = len(raw_query_map)
        if raw_total > 0:
            hit_raw = sum(1 for raw, qs in raw_query_map.items()
                          if qs & query_has_good)
            hit_rate = hit_raw / raw_total

    return {
        "query_total": qt,
        "有结果query数": qry_count,
        "有结果率": qry_count / qt if qt > 0 else 0,
        "有结果吸收率": qry_with_good / qry_count,
        "hit_rate": hit_rate,
        "产出率": yield_sum / qt if qt > 0 else 0,
        "Agood率": a_total / doc_total if doc_total > 0 else 0,
        "A+Bgood率": ab_total / doc_total if doc_total > 0 else 0,
        "query_Agood率": sum(qa_rates) / len(qa_rates),
        "query_A+Bgood率": sum(qab_rates) / len(qab_rates),
        "NDCG": sum(ndcg_list) / len(ndcg_list),
        "MAP": sum(map_list) / len(map_list) if map_list else 0,
        "MRR": sum(mrr_list) / len(mrr_list),
        "Timeliness": sum(year_scores) / len(year_scores) if year_scores else 0,
        "Authority": sum(auth_scores) / len(auth_scores) / 6 if auth_scores else 0,
    }


# ============================================================
# 工具
# ============================================================
def pipeline_name(a_path: Path) -> str:
    """A 表文件名 → 链路名（'_' 分隔取第一段）"""
    return a_path.stem.split("_")[0]


def build_pipelines(a_paths):
    pipelines = []
    for ap in a_paths:
        if not ap.exists():
            raise FileNotFoundError(f"A 表不存在: {ap}")
        pipelines.append((pipeline_name(ap), str(ap)))
    return pipelines


def build_scene_columns(mp, l1="一级场景", l2="二级场景", l3="三级场景"):
    """构建三个场景层级的合并展示名（下级为'/'时沿用上级）"""
    df = mp.copy()
    df[l2] = df.apply(lambda r: r[l1] if r[l2] == "/" else r[l2], axis=1)
    df[l3] = df.apply(lambda r: r[l2] if r[l3] == "/" else r[l3], axis=1)
    df["场景_一级"] = df[l1]
    df["场景_二级"] = df.apply(
        lambda r: r[l1] if r[l1] == r[l2] else f"{r[l1]}-{r[l2]}", axis=1)
    df["场景_三级"] = df.apply(
        lambda r: r["场景_二级"] if r[l2] == r[l3] else f"{r['场景_二级']}-{r[l3]}", axis=1)
    return df


# ============================================================
# 场景模式
# ============================================================
def run_scene_mode(pipelines, mapping_file, topk, output_dir, baseline_name):
    df_map = pd.read_excel(mapping_file)
    required = ["rawquery", "query", "一级场景", "二级场景", "三级场景"]
    missing = [c for c in required if c not in df_map.columns]
    if missing:
        raise KeyError(f"场景映射表缺少列: {missing}")

    print(f"[场景模式] 链路: {[n for n, _ in pipelines]}")
    print(f"  映射表: {mapping_file}")
    print(f"  权重: 硬编码（{len(SCENE_WEIGHTS)} 个二级场景）")
    print(f"  输出目录: {output_dir}")

    mp = build_scene_columns(df_map)
    scenes = {
        "一级场景": mp.groupby("场景_一级")["query"].apply(set).to_dict(),
        "二级场景": mp.groupby("场景_二级")["query"].apply(set).to_dict(),
        "三级场景": mp.groupby("场景_三级")["query"].apply(set).to_dict(),
    }
    # 每个场景对应的 rawquery → query 集合映射（用于计算 hit_rate）
    raw_query_maps = {
        "一级场景": {s: mp[mp["场景_一级"] == s].groupby("rawquery")["query"].apply(set).to_dict()
                     for s in scenes["一级场景"]},
        "二级场景": {s: mp[mp["场景_二级"] == s].groupby("rawquery")["query"].apply(set).to_dict()
                     for s in scenes["二级场景"]},
        "三级场景": {s: mp[mp["场景_三级"] == s].groupby("rawquery")["query"].apply(set).to_dict()
                     for s in scenes["三级场景"]},
    }

    all_rows = []
    for pname, ppath in pipelines:
        df_full = pd.read_excel(ppath)
        print(f"  {pname}: {len(df_full)} 行")
        pipeline_rows = []
        for level_name, scene_dict in scenes.items():
            for scene_name, qset in scene_dict.items():
                rq_map = raw_query_maps[level_name].get(scene_name, {})
                r = run_stats(df_full, qset, len(qset), topk, raw_query_map=rq_map)
                if r:
                    pipeline_rows.append({"链路": pname, "场景级别": level_name,
                                          "具体场景": scene_name, **r})

        # 加权总计（改为按二级场景加权）
        overall = {"链路": pname, "场景级别": OVERALL_NAME,
                   "具体场景": OVERALL_NAME, "query_total": mp["query"].nunique()}
        l2_rows = {r["具体场景"]: r for r in pipeline_rows if r["场景级别"] == "二级场景"}
        overall["有结果query数"] = sum(l2_rows[s].get("有结果query数", 0) for s in l2_rows)
        for m in METRIC_COLS:
            ws_sum = w_sum = 0.0
            for s, w in SCENE_WEIGHTS.items():
                if s in l2_rows:
                    ws_sum += l2_rows[s].get(m, 0) * w
                    w_sum += w
            overall[m] = ws_sum / w_sum if w_sum > 0 else 0
        pipeline_rows.append(overall)
        all_rows.extend(pipeline_rows)

    summary_df = pd.DataFrame(all_rows)
    LEVEL_ORDER = ["一级场景", "二级场景", "三级场景", OVERALL_NAME]
    PIPE_ORDER = [p[0] for p in pipelines]
    summary_df["链路"] = pd.Categorical(summary_df["链路"], categories=PIPE_ORDER, ordered=True)
    summary_df["场景级别"] = pd.Categorical(summary_df["场景级别"], categories=LEVEL_ORDER, ordered=True)
    summary_df = summary_df.sort_values(["链路", "场景级别", "具体场景"]).reset_index(drop=True)

    # 输出原始表 + 标色版
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    main_path = out_dir / "多级场景汇总.xlsx"
    color_path = out_dir / "多级场景汇总_最优最差.xlsx"
    summary_df.to_excel(main_path, index=False, engine="openpyxl")
    summary_df.to_excel(color_path, index=False, engine="openpyxl")

    # 标色（多链路最优最差 OR AB 对比）
    mode = "ab_compare" if baseline_name else "multi"
    color_targets = ({n for n in PIPE_ORDER if n != baseline_name}
                     if baseline_name else set(PIPE_ORDER))
    _scene_apply_color(summary_df, color_path, color_targets, mode, baseline_name)
    print(f"[OK] 多级场景汇总: {main_path}")
    print(f"[OK] 多级场景汇总（标色）: {color_path}")

    # 二级场景对比
    level2_path = out_dir / "二级场景关键指标纵向对比.xlsx"
    _gen_level2_compare(summary_df, color_path, level2_path)
    print(f"[OK] 二级场景对比: {level2_path}")

    # 总计加权紧凑对比
    overall_df = summary_df[summary_df["场景级别"] == OVERALL_NAME].copy()
    overall_df = overall_df.drop(columns=[c for c in ["场景级别", "具体场景", "有结果query数"]
                                          if c in overall_df.columns])
    overall_df["链路"] = pd.Categorical(overall_df["链路"], categories=PIPE_ORDER, ordered=True)
    overall_df = overall_df.sort_values("链路").reset_index(drop=True)
    overall_path = out_dir / "总计加权对比.xlsx"
    overall_df.to_excel(overall_path, index=False)
    _flat_apply_color(overall_path, baseline_name=baseline_name, pipe_order=PIPE_ORDER)
    print(f"[OK] 总计加权对比: {overall_path}")


def _scene_apply_color(summary_df, output_path, color_targets, mode, baseline_name):
    """场景模式 标色 + 数字格式 + 合并前两列"""
    # AB 对比基准查找表
    baseline_lookup = {}
    if mode == "ab_compare" and baseline_name:
        base_rows = summary_df[summary_df["链路"] == baseline_name]
        for _, row in base_rows.iterrows():
            for m in METRIC_COLS:
                if pd.notna(row.get(m)):
                    baseline_lookup[(row["场景级别"], row["具体场景"], m)] = row[m]
        print(f"  [标色] baseline={baseline_name}, 查找表 {len(baseline_lookup)} 条")

    best_map, worst_map = {}, {}
    if mode != "ab_compare":
        for (level, scene), sub in summary_df.groupby(["场景级别", "具体场景"], observed=True):
            if len(sub) < 2 or scene == DIFF_NAME:
                continue
            for m in METRIC_COLS:
                vals = sub[m].dropna()
                if len(vals) == 0:
                    continue
                bv, wv = vals.max(), vals.min()
                if bv == wv:
                    continue
                best_map[(level, scene, m)] = sub[sub[m] == bv]["链路"].astype(str).tolist()
                worst_map[(level, scene, m)] = sub[sub[m] == wv]["链路"].astype(str).tolist()

    wb = load_workbook(output_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    for idx, row_data in summary_df.iterrows():
        excel_row = idx + 2
        pipe = str(row_data.get("链路", ""))
        level = row_data.get("场景级别")
        scene = row_data.get("具体场景")
        is_overall = (level == OVERALL_NAME)

        for m in METRIC_COLS:
            if m not in col_idx:
                continue
            cell = ws.cell(row=excel_row, column=col_idx[m])
            if pipe in color_targets:
                if mode == "ab_compare":
                    test_val = row_data.get(m)
                    base_val = baseline_lookup.get((level, scene, m))
                    if pd.notna(test_val) and base_val is not None:
                        if test_val > base_val:
                            cell.fill = GREEN_FILL
                        elif test_val < base_val:
                            cell.fill = RED_FILL
                else:
                    best = best_map.get((level, scene, m), [])
                    worst = worst_map.get((level, scene, m), [])
                    if pipe in best:
                        cell.fill = GREEN_FILL
                    elif pipe in worst:
                        cell.fill = RED_FILL
                if is_overall:
                    cell.font = BOLD_FONT
            elif is_overall:
                cell.font = BOLD_FONT

    # 数字格式
    for r in range(2, ws.max_row + 1):
        for c in range(6, 17):
            ws.cell(row=r, column=c).number_format = "0.00%"
        for c in [17, 18]:
            ws.cell(row=r, column=c).number_format = "0.0000"

    # 列宽
    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J",
                       "K", "L", "M", "N", "O", "P", "Q", "R"]:
        max_len = 0
        col_n = ord(col_letter) - ord("A") + 1
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col_n).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 28)

    # 合并前两列
    for col_n in [1, 2]:
        start_row = 2
        prev = ws.cell(row=start_row, column=col_n).value
        for row in range(3, ws.max_row + 2):
            cur = ws.cell(row=row, column=col_n).value if row <= ws.max_row else None
            if cur != prev:
                if row - 1 > start_row:
                    ws.merge_cells(start_row=start_row, start_column=col_n,
                                   end_row=row - 1, end_column=col_n)
                    ws.cell(row=start_row, column=col_n).alignment = Alignment(
                        horizontal="left", vertical="center")
                start_row = row
                prev = cur

    wb.save(output_path)


def _gen_level2_compare(summary_df, color_file_path, output_path):
    """生成二级场景关键指标纵向对比表（带标色复制）"""
    df = pd.read_excel(color_file_path)
    df["链路"] = df["链路"].ffill()
    df["场景级别"] = df["场景级别"].ffill()
    df_l2 = df[df["场景级别"].str.contains("二级", na=False)].copy()

    wb_color = load_workbook(color_file_path)
    ws_color = wb_color.active
    scene_row_map = {}
    for idx, row in df_l2.iterrows():
        scene_row_map.setdefault(row["具体场景"], {})[row["链路"]] = idx + 2

    col_map = {}
    for ci, cell in enumerate(ws_color[1], 1):
        if cell.value:
            col_map[cell.value] = ci

    wb = Workbook()
    ws = wb.active
    key_metrics = ["有结果率", "Agood率", "A+Bgood率", "Timeliness", "Authority"]
    pipelines = sorted(df_l2["链路"].unique())
    headers = ["一级场景", "二级场景"]
    for m in key_metrics:
        for p in pipelines:
            headers.append(f"{p}-{m}")
    ws.append(headers)

    scenes = sorted(df_l2["具体场景"].unique())
    prev_l1 = None
    nri = 2
    for scene in scenes:
        l1 = scene.split("-")[0] if "-" in scene else scene
        if l1 == prev_l1:
            ws.cell(row=nri, column=1).value = ""
        else:
            ws.cell(row=nri, column=1).value = l1
            prev_l1 = l1
        ws.cell(row=nri, column=2).value = scene

        nci = 3
        for m in key_metrics:
            for p in pipelines:
                sub = df_l2[(df_l2["具体场景"] == scene) & (df_l2["链路"] == p)]
                if len(sub) > 0 and m in col_map and scene in scene_row_map and p in scene_row_map[scene]:
                    ws.cell(row=nri, column=nci).value = sub[m].iloc[0]
                    orig = ws_color.cell(row=scene_row_map[scene][p], column=col_map[m])
                    if orig.fill and orig.fill.start_color and orig.fill.start_color.rgb:
                        rgb = orig.fill.start_color.rgb
                        if rgb and len(str(rgb)) >= 6:
                            ws.cell(row=nri, column=nci).fill = PatternFill(
                                start_color=rgb, end_color=rgb, fill_type="solid")
                nci += 1
        nri += 1

    # 数字格式
    for r in range(2, ws.max_row + 1):
        for c in range(3, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                header = ws.cell(row=1, column=c).value or ""
                if "Timeliness" in header or "Authority" in header:
                    cell.number_format = "0.0000"
                else:
                    cell.number_format = "0.00%"

    # 列宽
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 24)

    wb.save(output_path)


def _flat_apply_color(xlsx_path, baseline_name, pipe_order):
    """对紧凑单层表（每行一条链路）标色"""
    HIGHER_BETTER = {"有结果率", "有结果吸收率", "hit_rate", "产出率", "Agood率", "A+Bgood率",
                     "query_Agood率", "query_A+Bgood率",
                     "NDCG", "MAP", "MRR", "Timeliness", "Authority"}
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    if "链路" not in col_idx:
        wb.save(xlsx_path); return
    name_col = col_idx["链路"]

    baseline_row, base_vals = None, {}
    if baseline_name:
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=name_col).value == baseline_name:
                baseline_row = r
                for h in headers:
                    base_vals[h] = ws.cell(row=r, column=col_idx[h]).value
                break

    if baseline_row is not None:
        for r in range(2, ws.max_row + 1):
            if r == baseline_row:
                continue
            for h in headers:
                if h == "链路" or h not in HIGHER_BETTER:
                    continue
                cell = ws.cell(row=r, column=col_idx[h])
                try:
                    v = float(cell.value); base = float(base_vals.get(h))
                except (TypeError, ValueError):
                    continue
                if v > base:
                    cell.fill = GREEN_FILL
                elif v < base:
                    cell.fill = RED_FILL
        for h in headers:
            ws.cell(row=baseline_row, column=col_idx[h]).font = BOLD_FONT
    else:
        for h in headers:
            if h == "链路" or h not in HIGHER_BETTER:
                continue
            vals = []
            for r in range(2, ws.max_row + 1):
                try:
                    vals.append((r, float(ws.cell(row=r, column=col_idx[h]).value)))
                except (TypeError, ValueError):
                    pass
            if len(vals) < 2:
                continue
            best_r = max(vals, key=lambda x: x[1])[0]
            worst_r = min(vals, key=lambda x: x[1])[0]
            ws.cell(row=best_r, column=col_idx[h]).fill = GREEN_FILL
            ws.cell(row=worst_r, column=col_idx[h]).fill = RED_FILL

    for r in range(2, ws.max_row + 1):
        for h in headers:
            cell = ws.cell(row=r, column=col_idx[h])
            if h == "链路" or not isinstance(cell.value, (int, float)):
                continue
            if h in {"Timeliness", "Authority"}:
                cell.number_format = "0.0000"
            elif h.endswith("率") or h in {"NDCG", "MAP", "MRR"}:
                cell.number_format = "0.00%"

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 22)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
    wb.save(xlsx_path)


# ============================================================
# 类型 / 普通总计 模式
# ============================================================
def run_type_or_overall_mode(pipelines, mapping_file, topk, output_dir, mode_name,
                              baseline_name=None):
    """两种模式共用：构建 dimensions（"类型"按类型列分组；"普通总计"只跑全量），
    然后按 (链路, 维度) 跑统计，输出汇总表。"""
    df_map = pd.read_excel(mapping_file)
    if mode_name == "类型":
        required = ["rawquery", "query", "类型"]
        missing = [c for c in required if c not in df_map.columns]
        if missing:
            raise KeyError(f"类型映射表缺少列: {missing}")
        types = sorted(df_map["类型"].dropna().astype(str).unique())
        dimensions = {"全量": None}
        for t in types:
            dimensions[t] = [t]
        type_col = "类型"
    else:  # 普通总计
        required = ["rawquery", "query"]
        missing = [c for c in required if c not in df_map.columns]
        if missing:
            raise KeyError(f"映射表缺少列: {missing}")
        dimensions = {"全量": None}
        type_col = None  # 不需要

    print(f"[{mode_name}模式] 链路: {[n for n, _ in pipelines]}")
    print(f"  映射表: {mapping_file}")
    print(f"  维度: {list(dimensions.keys())}")
    print(f"  输出目录: {output_dir}")

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    all_results = []
    for pname, ppath in pipelines:
        df = pd.read_excel(ppath)
        print(f"  -- {pname}：{len(df)} 行 --")
        for dim_name, type_filter in dimensions.items():
            if type_filter is None:
                qset = set(df_map["query"].unique())
                rq_map = df_map.groupby("rawquery")["query"].apply(set).to_dict()
            else:
                sub_map = df_map[df_map[type_col].isin(type_filter)]
                qset = set(sub_map["query"].unique())
                rq_map = sub_map.groupby("rawquery")["query"].apply(set).to_dict()
            qt = len(qset)
            if qt == 0:
                print(f"    [跳过] {dim_name}: query 集合为空")
                continue
            r = run_stats(df, qset, qt, topk, raw_query_map=rq_map)
            if r:
                row = {"链路": pname, "维度": dim_name, **r}
                all_results.append(row)
                print(f"    {dim_name}: 有结果率={r['有结果率']:.4f} | "
                      f"hit_rate={r['hit_rate']:.4f} | "
                      f"A+Bgood率={r['A+Bgood率']:.4f} | "
                      f"NDCG={r['NDCG']:.4f}")

    summary = pd.DataFrame(all_results)
    PIPE_ORDER = [p[0] for p in pipelines]
    DIM_ORDER = list(dimensions.keys())
    summary["链路"] = pd.Categorical(summary["链路"], categories=PIPE_ORDER, ordered=True)
    summary["维度"] = pd.Categorical(summary["维度"], categories=DIM_ORDER, ordered=True)
    summary = summary.sort_values(["链路", "维度"]).reset_index(drop=True)

    out_name = f"汇总统计表_{len(pipelines)}链路{len(dimensions)}维度_topK{topk}.xlsx"
    out_path = out_dir / out_name
    summary.to_excel(out_path, index=False, engine="openpyxl")

    # 标色：每个 (维度, 指标) 找最优最差，或与 baseline 对比
    _multi_dim_apply_color(out_path, DIM_ORDER, baseline_name, PIPE_ORDER)
    print(f"[OK] 汇总表: {out_path}")


def _multi_dim_apply_color(xlsx_path, dim_order, baseline_name, pipe_order):
    """多维度汇总表（链路 × 维度 × 指标）标色"""
    HIGHER_BETTER = {"有结果率", "有结果吸收率", "hit_rate", "产出率", "Agood率", "A+Bgood率",
                     "query_Agood率", "query_A+Bgood率",
                     "NDCG", "MAP", "MRR", "Timeliness", "Authority"}
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    if "链路" not in col_idx or "维度" not in col_idx:
        wb.save(xlsx_path); return

    df = pd.read_excel(xlsx_path)
    metric_cols = [h for h in headers if h in HIGHER_BETTER]

    if baseline_name:
        # AB 对比模式：每个 (维度, 指标) 与 baseline 对比
        for dim in dim_order:
            sub = df[df["维度"] == dim]
            if len(sub) == 0:
                continue
            base_row = sub[sub["链路"] == baseline_name]
            if len(base_row) == 0:
                continue
            for m in metric_cols:
                if m not in col_idx:
                    continue
                base_v = base_row[m].iloc[0]
                if pd.isna(base_v):
                    continue
                for _, row in sub.iterrows():
                    pipe = str(row["链路"])
                    if pipe == baseline_name:
                        continue
                    v = row[m]
                    if pd.isna(v):
                        continue
                    df_idx = row.name
                    excel_row = df_idx + 2
                    cell = ws.cell(row=excel_row, column=col_idx[m])
                    if v > base_v:
                        cell.fill = GREEN_FILL
                    elif v < base_v:
                        cell.fill = RED_FILL
    else:
        # 多链路模式：每列最优最差
        for dim in dim_order:
            sub = df[df["维度"] == dim]
            if len(sub) < 2:
                continue
            for m in metric_cols:
                if m not in col_idx:
                    continue
                vals = sub[m].dropna()
                if len(vals) == 0:
                    continue
                bv, wv = vals.max(), vals.min()
                if bv == wv:
                    continue
                for _, row in sub.iterrows():
                    excel_row = row.name + 2
                    cell = ws.cell(row=excel_row, column=col_idx[m])
                    if row[m] == bv:
                        cell.fill = GREEN_FILL
                    elif row[m] == wv:
                        cell.fill = RED_FILL

    # baseline 行加粗
    if baseline_name:
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=col_idx["链路"]).value) == baseline_name:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).font = BOLD_FONT

    # 数字格式
    for r in range(2, ws.max_row + 1):
        for h in headers:
            cell = ws.cell(row=r, column=col_idx[h])
            if h in {"链路", "维度"} or not isinstance(cell.value, (int, float)):
                continue
            if h in {"Timeliness", "Authority"}:
                cell.number_format = "0.0000"
            elif h.endswith("率") or h in {"NDCG", "MAP", "MRR"}:
                cell.number_format = "0.00%"

    # 列宽
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 22)

    wb.save(xlsx_path)


# ============================================================
# 主入口
# ============================================================
def run(stats_mode, a_table_paths, mapping_file, output_dir=None,
        topk=3, baseline_name=None) -> int:
    a_paths = [Path(p).resolve() for p in a_table_paths]
    mapping_file = Path(mapping_file).resolve()
    if not mapping_file.exists():
        print(f"[ERR] 映射表不存在: {mapping_file}"); return 1

    if output_dir is None:
        output_dir = a_paths[0].parent / f"统计结果_{stats_mode}"
    output_dir = Path(output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    pipelines = build_pipelines(a_paths)
    print("=" * 70)
    print(f"统计模式: {stats_mode}")
    print(f"链路数: {len(pipelines)} | topK: {topk}")
    if baseline_name:
        print(f"baseline: {baseline_name}（AB 对比模式）")
    else:
        print("多链路模式（每列最优绿底/最差红底）")
    print("=" * 70)

    if stats_mode == "场景":
        run_scene_mode(pipelines, mapping_file, topk, output_dir, baseline_name)
    elif stats_mode == "类型":
        run_type_or_overall_mode(pipelines, mapping_file, topk, output_dir, "类型", baseline_name)
    elif stats_mode == "普通总计":
        run_type_or_overall_mode(pipelines, mapping_file, topk, output_dir, "普通总计", baseline_name)
    else:
        print(f"[ERR] 未知 STATS_MODE: {stats_mode}（应选 '场景' / '类型' / '普通总计'）")
        return 1

    print(f"\n[ALL DONE] 结果已输出到: {output_dir}")
    return 0


# ============================================================
# 配置区：修改下面的入参后直接运行本脚本
# ============================================================
if __name__ == "__main__":
    # ----- ① 统计模式：三选一 -----
    # "场景"      : 按一/二/三级场景统计 + 加权总计；输出 3 张表
    #              （映射表需有: rawquery, query, 一级场景, 二级场景, 三级场景）
    # "类型"      : 按各类型 + 总计统计；输出 1 张总表
    #              （映射表需有: rawquery, query, 类型）
    # "普通总计"  : 只看总计；输出 1 张总表
    #              （映射表需有: rawquery, query）
    STATS_MODE = "场景"
    # ----- ② N 个 A 表（已回填标注的 *_AB合并表.xlsx） -----
    A_TABLES = [
        r"D:\医学检索test\端到端测试\标注库AB\结果统计\1235更新链路全量测试集_合肥_0617.xlsx",
        r"D:\医学检索test\端到端测试\标注库AB\结果统计\1236全量测试集跑测_合肥_0618.xlsx",
        r"D:\医学检索test\端到端测试\标注库AB\结果统计\线上权威全量测试集_合肥_0617.xlsx"
    ]
    # ----- ③ 映射表 -----
    MAPPING_FILE = r"D:\医学检索test\端到端测试\2026年\测试集1584映射表(1).xlsx"
    # ----- ④ 输出目录 -----
    # None = 自动选第一个 A 表的同级目录下"统计结果_<模式>"
    OUTPUT_DIR = None
    # ----- ⑤ topK -----
    TOPK = 3
    # ----- ⑥ baseline 链路名（可选） -----
    # None         = 多链路模式（每列最优绿底、最差红底）
    # "线上版本测试集" = AB 对比模式（其他链路 vs baseline，>baseline 绿底、<baseline 红底）
    BASELINE_NAME = "线上权威全量测试集"
    # ============================================================
    # 以下不用改
    # ============================================================
    sys.exit(run(STATS_MODE, A_TABLES, MAPPING_FILE, OUTPUT_DIR, TOPK, BASELINE_NAME))
