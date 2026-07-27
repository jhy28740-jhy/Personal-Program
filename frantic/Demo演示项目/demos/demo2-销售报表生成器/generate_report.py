"""
销售数据自动报表生成器 - Demo版
功能：从原始CSV生成带图表的Excel销售月报
作者：数据服务工作室
"""

import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os


class SalesReportGenerator:
    def __init__(self, input_file='原始数据.csv'):
        """初始化报表生成器"""
        self.input_file = input_file
        self.output_file = f'销售月报_{datetime.now().strftime("%Y%m")}.xlsx'

    def load_data(self):
        """加载原始数据"""
        print(f"正在读取数据: {self.input_file}")

        if not os.path.exists(self.input_file):
            print("[!] 未找到原始数据，使用Demo示例数据")
            self.create_sample_data()

        df = pd.read_csv(self.input_file, encoding='utf-8-sig')
        print(f"[OK] 成功加载 {len(df)} 条销售记录")
        return df

    def create_sample_data(self):
        """创建示例数据"""
        import random

        products = ['iPhone 15', 'MacBook Pro', 'AirPods Pro', 'iPad Air', 'Apple Watch']
        customers = ['阿里巴巴', '腾讯', '字节跳动', '京东', '百度', '美团', '小米', '华为']
        regions = ['华东', '华北', '华南', '西南', '东北']

        data = []
        for i in range(100):
            data.append({
                '订单日期': f'2024-07-{random.randint(1, 28):02d}',
                '客户名称': random.choice(customers),
                '产品名称': random.choice(products),
                '销售数量': random.randint(1, 50),
                '单价': random.choice([5999, 12999, 1999, 4999, 2999]),
                '地区': random.choice(regions)
            })

        df = pd.DataFrame(data)
        df['销售额'] = df['销售数量'] * df['单价']
        df.to_csv(self.input_file, index=False, encoding='utf-8-sig')
        print(f"[OK] 已创建示例数据: {self.input_file}")

    def analyze_data(self, df):
        """数据分析"""
        print("正在分析数据...")

        # 计算销售额
        if '销售额' not in df.columns:
            df['销售额'] = df['销售数量'] * df['单价']

        # 按产品汇总
        product_summary = df.groupby('产品名称').agg({
            '销售数量': 'sum',
            '销售额': 'sum'
        }).sort_values('销售额', ascending=False).reset_index()

        # 按客户汇总
        customer_summary = df.groupby('客户名称').agg({
            '销售额': 'sum'
        }).sort_values('销售额', ascending=False).head(10).reset_index()

        # 按地区汇总
        region_summary = df.groupby('地区').agg({
            '销售额': 'sum',
            '销售数量': 'sum'
        }).reset_index()

        # 总体统计
        total_sales = df['销售额'].sum()
        total_quantity = df['销售数量'].sum()
        avg_price = df['单价'].mean()
        order_count = len(df)

        summary = {
            'product': product_summary,
            'customer': customer_summary,
            'region': region_summary,
            'total_sales': total_sales,
            'total_quantity': total_quantity,
            'avg_price': avg_price,
            'order_count': order_count
        }

        print(f"[OK] 分析完成: 总销售额 RMB{total_sales:,.0f}, 订单数 {order_count}")
        return summary

    def create_excel_report(self, df, summary):
        """生成Excel报表"""
        print(f"正在生成报表: {self.output_file}")

        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            # Sheet 1: 总览
            self.create_overview_sheet(writer, summary)

            # Sheet 2: 产品分析
            summary['product'].to_excel(writer, sheet_name='产品分析', index=False)

            # Sheet 3: 客户分析
            summary['customer'].to_excel(writer, sheet_name='客户分析', index=False)

            # Sheet 4: 地区分析
            summary['region'].to_excel(writer, sheet_name='地区分析', index=False)

            # Sheet 5: 原始数据
            df.to_excel(writer, sheet_name='原始数据', index=False)

        # 添加图表
        self.add_charts(summary)

        print(f"[OK] 报表生成完成: {self.output_file}")

    def create_overview_sheet(self, writer, summary):
        """创建总览页"""
        overview_data = {
            '指标': ['总销售额', '总销售数量', '平均单价', '订单总数'],
            '数值': [
                f"RMB{summary['total_sales']:,.2f}",
                f"{summary['total_quantity']:,}",
                f"RMB{summary['avg_price']:,.2f}",
                f"{summary['order_count']:,}"
            ]
        }
        df_overview = pd.DataFrame(overview_data)
        df_overview.to_excel(writer, sheet_name='总览', index=False)

    def add_charts(self, summary):
        """添加图表"""
        wb = openpyxl.load_workbook(self.output_file)

        # 产品销售额柱状图
        if '产品分析' in wb.sheetnames:
            ws = wb['产品分析']
            chart = BarChart()
            chart.title = "产品销售额对比"
            chart.x_axis.title = "产品"
            chart.y_axis.title = "销售额 (元)"

            data = Reference(ws, min_col=3, min_row=1, max_row=len(summary['product'])+1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(summary['product'])+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "E2")

        # 地区销售额饼图
        if '地区分析' in wb.sheetnames:
            ws = wb['地区分析']
            chart = PieChart()
            chart.title = "地区销售额占比"

            data = Reference(ws, min_col=2, min_row=1, max_row=len(summary['region'])+1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(summary['region'])+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "E2")

        wb.save(self.output_file)
        print("[OK] 图表已添加")

    def generate(self):
        """执行报表生成"""
        print("=" * 50)
        print("销售数据自动报表生成器 v1.0")
        print("=" * 50)

        # 1. 加载数据
        df = self.load_data()

        # 2. 分析数据
        summary = self.analyze_data(df)

        # 3. 生成Excel报表
        self.create_excel_report(df, summary)

        print("\n" + "=" * 50)
        print("[OK] 报表生成完成！")
        print(f" 查看报表: {self.output_file}")
        print("=" * 50)

        return self.output_file


if __name__ == "__main__":
    generator = SalesReportGenerator()
    generator.generate()
